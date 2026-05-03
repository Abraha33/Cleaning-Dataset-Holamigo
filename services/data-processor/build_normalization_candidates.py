"""
Genera CSV de normalización (candidatos) sin modificar product_master.
No inventa mappings: todo lo dudoso queda status=propuesto con notas.
"""
from __future__ import annotations

import ast
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "normalization_output"

PATHS = {
    "brands_json": ROOT / "brands.json",
    "marcas_oraculo": ROOT / "marcas_oraculo.json",
    "categories_tienda": ROOT / "Categories_Tienda_Linea.xlsx",
    "brands_xlsx": ROOT / "brands.xlsx",
    "ecoplast": ROOT / "ecoplast_data.json",
    "colors": ROOT / "colors.json",
    "material": ROOT / "material.json",
    "unities": ROOT / "unities.json",
    "categorias_internal": ROOT / "categorias.json",
    "categorias_oraculo": ROOT / "categorias_oraculo.json",
}


def load_json(path: Path):
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def parse_categorias_json(path: Path) -> list[dict]:
    raw = load_json(path)
    if not raw:
        return []
    rows = []
    for item in raw:
        if not isinstance(item, str):
            continue
        s = item.strip()
        if " - Generic -" in s:
            s = s.split(" - Generic -")[0].rstrip()
            if not s.endswith("}"):
                s = s + "}"
        try:
            d = ast.literal_eval(s)
            if isinstance(d, dict):
                rows.append(d)
        except (SyntaxError, ValueError):
            rows.append({"parse_error": True, "raw": item[:200]})
    return rows


def try_load_xlsx_categories(path: Path) -> tuple[list[dict], str | None]:
    if not path.exists():
        return [], f"Archivo no encontrado: {path}"
    try:
        import openpyxl
    except ImportError:
        return [], "openpyxl no instalado; pip install openpyxl"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    out = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        out.append(dict(zip(headers, row)))
    wb.close()
    return out, None


def try_load_xlsx_brands(path: Path) -> tuple[list[dict], str | None]:
    return try_load_xlsx_categories(path)


def norm_key(s: str) -> str:
    return " ".join(s.lower().strip().split())


def write_csv(name: str, fieldnames: list[str], rows: list[dict]):
    OUT.mkdir(parents=True, exist_ok=True)
    p = OUT / name
    with p.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    return p


def build_brands(
    brands_list: list,
    oraculo_list: list,
    brands_xlsx_rows: list[dict] | None,
    ambiguos: list[str],
) -> tuple[list[dict], list[dict]]:
    """Clusters by casefold; canonical picks first brands.json then oraculo."""
    cluster_strings: dict[str, set[str]] = defaultdict(set)
    for b in brands_list or []:
        cluster_strings[norm_key(b)].add(b)
    for b in oraculo_list or []:
        cluster_strings[norm_key(b)].add(b)
    xlsx_strings = []
    if brands_xlsx_rows:
        for r in brands_xlsx_rows:
            for k in ("marca", "brand", "nombre", "Nombre", "MARCA"):
                if k in r and r[k]:
                    xlsx_strings.append(str(r[k]).strip())
                    break
        for b in xlsx_strings:
            cluster_strings[norm_key(b)].add(b)

    # Prefer canonical: from brands.json exact if any in cluster starts with brands list preference
    brands_set = {str(x) for x in (brands_list or [])}
    master = []
    aliases = []
    bid = 0
    aid = 0
    for nk, variants in sorted(cluster_strings.items(), key=lambda x: x[0]):
        if not nk:
            continue
        variants_list = sorted(variants)
        chosen = None
        for v in variants_list:
            if v in brands_set:
                chosen = v
                break
        if chosen is None:
            chosen = variants_list[0]
        bid += 1
        brand_id = f"BM{bid:05d}"
        master.append(
            {
                "brand_id": brand_id,
                "canonical_name": chosen,
                "norm_key": nk,
                "sources_present": "|".join(
                    sorted(
                        {
                            *(["brands_json"] if any(v in brands_set for v in variants_list) else []),
                            *(
                                ["marcas_oraculo"]
                                if oraculo_list and any(v in set(oraculo_list) for v in variants_list)
                                else []
                            ),
                            *(
                                ["brands_xlsx"]
                                if xlsx_strings and any(v in set(xlsx_strings) for v in variants_list)
                                else []
                            ),
                        }
                    )
                ),
                "variant_count": len(variants_list),
                "status": "propuesto",
                "notas": "Agrupación por igualdad sin distinguir mayúsculas/espacios; requiere validación humana.",
            }
        )
        if len(variants_list) > 1:
            ambiguos.append(
                f"Marca cluster '{chosen}': {len(variants_list)} variantes ortográficas — revisar si son la misma marca."
            )
        for v in variants_list:
            if v == chosen:
                continue
            aid += 1
            same_chars = v == chosen
            aliases.append(
                {
                    "alias_id": f"BA{aid:05d}",
                    "alias_raw": v,
                    "brand_id": brand_id,
                    "canonical_name": chosen,
                    "match_type": "exact_literal" if same_chars else "case_or_spacing_fold",
                    "status": "aprobado" if same_chars else "propuesto",
                    "notas": ""
                    if same_chars
                    else "Misma clave normalizada (minúsculas/espacios) que el canónico elegido; no se asume equivalencia de marca sin revisión.",
                }
            )
    despace_to_canonicals: dict[str, list[str]] = defaultdict(list)
    for row in master:
        dk = norm_key(row["canonical_name"]).replace(" ", "")
        despace_to_canonicals[dk].append(row["canonical_name"])
    for dk, names in despace_to_canonicals.items():
        uniq = sorted(set(names))
        if len(uniq) > 1:
            ambiguos.append(
                "Posibles marcas duplicadas (mismo texto sin espacios; NO fusionadas): "
                + " | ".join(uniq)
            )
    return master, aliases


def build_normalization_dict(colors, material, unities, ambiguos: list[str]) -> list[dict]:
    rows = []
    nid = 0

    def add_row(domain, raw, canonical_key, canonical_display, status, notas):
        nonlocal nid
        nid += 1
        rows.append(
            {
                "norm_id": f"ND{nid:05d}",
                "domain": domain,
                "raw_variant": raw,
                "canonical_key": canonical_key,
                "canonical_display": canonical_display,
                "status": status,
                "notas": notas,
            }
        )

    for c in colors or []:
        cid = c.get("id", "")
        nombre = c.get("nombre", "")
        abbr = c.get("abreviatura", "")
        add_row(
            "color",
            nombre,
            cid,
            nombre,
            "aprobado",
            "Match exacto con catálogo colors.json (nombre).",
        )
        if abbr and abbr != nombre:
            add_row(
                "color",
                abbr,
                cid,
                nombre,
                "aprobado",
                "Match exacto con catálogo colors.json (abreviatura).",
            )

    for m in material or []:
        mid = m.get("id", "")
        nombre = m.get("nombre", "")
        abbr = m.get("abreviatura", "")
        add_row(
            "material",
            nombre,
            mid,
            nombre,
            "aprobado",
            "Match exacto con catálogo material.json (nombre).",
        )
        if abbr and abbr != nombre:
            add_row(
                "material",
                abbr,
                mid,
                nombre,
                "aprobado",
                "Match exacto con catálogo material.json (abreviatura).",
            )

    for u in unities or []:
        uid = u.get("id", "")
        nombre = u.get("nombre", "")
        abbr = u.get("abreviatura", "")
        tipo = u.get("tipo", "")
        add_row(
            "unit",
            nombre,
            uid,
            nombre,
            "aprobado",
            f"Catálogo unities.json (tipo={tipo}).",
        )
        if abbr and abbr != nombre:
            add_row(
                "unit",
                abbr,
                uid,
                nombre,
                "aprobado",
                f"Catálogo unities.json abreviatura (tipo={tipo}).",
            )

    return rows


def build_category_bridge(
    ecoplast_categories: set[str],
    oracle_paths: set[str],
    official_rows: list[dict],
    internal_path_to_code: dict[str, str],
    ambiguos: list[str],
) -> list[dict]:
    rows = []
    oid = 0
    if official_rows and PATHS["categories_tienda"].exists():
        ambiguos.append(
            "Categories_Tienda_Linea.xlsx está presente pero no hay reglas de columnas acordadas: "
            "suggested_official_* queda vacío; completar manualmente o ampliar el script."
        )
    for cat in sorted(ecoplast_categories):
        oid += 1
        exact_oracle = cat in oracle_paths
        nk = norm_key(cat)
        internal_code = internal_path_to_code.get(nk, "")
        suggested_codigo = ""
        suggested_path = ""
        status = "propuesto"
        notas = []
        if not PATHS["categories_tienda"].exists():
            notas.append("Categories_Tienda_Linea.xlsx no está en el workspace; no hay objetivo oficial automático.")
        if exact_oracle:
            notas.append(
                "El texto coincide exactamente con una ruta en categorias_oraculo.json "
                "(referencia auxiliar; no sustituye el árbol oficial en Excel)."
            )
        else:
            notas.append("No hay coincidencia textual exacta con categorias_oraculo.json; mapping pendiente.")
        if internal_code:
            notas.append(
                f"La ruta normalizada coincide con el árbol armado desde categorias.json (codigo nodo={internal_code}); "
                "sigue siendo propuesto hasta cruzar con Categories_Tienda_Linea."
            )
        match_method = "none"
        if exact_oracle and internal_code:
            match_method = "exact_text_oracle+internal_path_norm_exact"
        elif exact_oracle:
            match_method = "exact_text_oracle"
        elif internal_code:
            match_method = "internal_path_norm_exact"
        rows.append(
            {
                "bridge_id": f"CB{oid:05d}",
                "source_system": "ecoplast_data",
                "source_category_path": cat,
                "suggested_official_codigo": suggested_codigo,
                "suggested_official_path": suggested_path,
                "oracle_path_exact_match": cat if exact_oracle else "",
                "internal_erp_codigo_hint": internal_code or "",
                "match_method": match_method,
                "status": status,
                "notas": " ".join(notas),
            }
        )
    return rows


def internal_paths_by_norm_key(internal_rows: list[dict]) -> dict[str, str]:
    """
    Construye ruta 'Padre - Hijo - ...' desde categorias.json y devuelve
    norm_key(ruta_completa) -> codigo del nodo hoja (mismo codigo del nodo).
    """
    by_code: dict[str, dict] = {}
    for r in internal_rows:
        if r.get("parse_error"):
            continue
        c = r.get("codigo")
        if c:
            by_code[str(c)] = r

    def path_for(codigo: str | None) -> str:
        parts: list[str] = []
        seen: set[str] = set()
        while codigo and codigo not in seen:
            seen.add(codigo)
            node = by_code.get(codigo)
            if not node:
                break
            n = node.get("nombre")
            if n:
                parts.append(str(n))
            codigo = node.get("codigo_categoria_padre")
        return " - ".join(reversed(parts))

    out: dict[str, str] = {}
    for code in by_code:
        full = path_for(code)
        if full:
            out[norm_key(full)] = code
    return out


PRESENTATION_PATTERNS = [
    (re.compile(r"(?i)caja\s*x\s*(\d+)"), "caja_x_n"),
    (re.compile(r"(?i)(\d+)\s*und"), "n_und"),
    (re.compile(r"(?i)millar"), "millar"),
    (re.compile(r"(?i)a\s*partir\s*de\s*(\d+)"), "tier_min_qty"),
]


def build_unit_candidates(unities, ecoplast_presentations: set[str], ambiguos: list[str]) -> list[dict]:
    rows = []
    uid = 0
    uby_abbr = {u.get("abreviatura", "").lower(): u for u in (unities or [])}

    for u in unities or []:
        uid += 1
        rows.append(
            {
                "unit_candidate_id": f"UC{uid:05d}",
                "raw_presentation": u.get("nombre", ""),
                "source": "unities_json",
                "parsed_quantity": "",
                "parsed_unit_token": u.get("abreviatura", ""),
                "suggested_unity_id": u.get("id", ""),
                "factor_to_base": "",
                "status": "aprobado",
                "notas": "Registro maestro unities.json.",
            }
        )

    und_like = 0
    for pres in sorted(ecoplast_presentations):
        if not pres or not str(pres).strip():
            continue
        pl = str(pres).strip()
        uid += 1
        parsed_qty = ""
        method = "ecoplast_variant"
        factor = ""
        for rx, _tag in PRESENTATION_PATTERNS:
            m = rx.search(pl)
            if m:
                if m.lastindex:
                    parsed_qty = m.group(1)
                break
        sugg_id = ""
        low = pl.lower()
        if re.search(r"\bund\b|unidad", low):
            und_like += 1
        for abbr, u in uby_abbr.items():
            if abbr and re.search(rf"\b{re.escape(abbr)}\b", low):
                sugg_id = u.get("id", "")
        rows.append(
            {
                "unit_candidate_id": f"UC{uid:05d}",
                "raw_presentation": pl,
                "source": method,
                "parsed_quantity": parsed_qty,
                "parsed_unit_token": "",
                "suggested_unity_id": sugg_id,
                "factor_to_base": factor,
                "status": "propuesto",
                "notas": "Texto libre desde ecoplast; cantidad/factor solo si patrón regex coincidió; validar humano.",
            }
        )
    if und_like:
        ambiguos.append(
            f"{und_like} presentaciones ecoplast contienen 'und'/'unidad'; no se mapean a unidades del maestro sin regla aprobada."
        )
    return rows


def main():
    ambiguos: list[str] = []
    warnings: list[str] = []

    brands_list = load_json(PATHS["brands_json"]) or []
    oraculo_list = load_json(PATHS["marcas_oraculo"]) or []
    colors = load_json(PATHS["colors"])
    material = load_json(PATHS["material"])
    unities = load_json(PATHS["unities"])
    oracle_cats = load_json(PATHS["categorias_oraculo"]) or []
    oracle_paths = set(oracle_cats) if isinstance(oracle_cats, list) else set()

    internal = parse_categorias_json(PATHS["categorias_internal"])
    internal_path_to_code = internal_paths_by_norm_key(internal)
    parse_errors = sum(1 for r in internal if r.get("parse_error"))
    if parse_errors:
        warnings.append(f"categorias.json: {parse_errors} líneas no parseadas.")

    official_rows, err = try_load_xlsx_categories(PATHS["categories_tienda"])
    if err:
        warnings.append(err)

    brands_xlsx, bx_err = try_load_xlsx_brands(PATHS["brands_xlsx"])
    if bx_err:
        warnings.append(bx_err)
    elif brands_xlsx:
        warnings.append(f"brands.xlsx cargado: {len(brands_xlsx)} filas (columna marca inferida).")

    ecoplast = load_json(PATHS["ecoplast"]) or []
    eco_cats: set[str] = set()
    eco_pres: set[str] = set()
    for item in ecoplast:
        c = item.get("category")
        if c:
            eco_cats.add(str(c).strip())
        for v in item.get("variants") or []:
            p = v.get("presentation")
            if p:
                eco_pres.add(str(p).strip())

    brand_master, brand_alias = build_brands(
        brands_list, oraculo_list, brands_xlsx if brands_xlsx else None, ambiguos
    )
    norm_dict = build_normalization_dict(colors, material, unities, ambiguos)
    cat_bridge = build_category_bridge(eco_cats, oracle_paths, official_rows, internal_path_to_code, ambiguos)
    unit_rows = build_unit_candidates(unities, eco_pres, ambiguos)

    write_csv(
        "brand_master.csv",
        [
            "brand_id",
            "canonical_name",
            "norm_key",
            "sources_present",
            "variant_count",
            "status",
            "notas",
        ],
        brand_master,
    )
    write_csv(
        "brand_alias.csv",
        [
            "alias_id",
            "alias_raw",
            "brand_id",
            "canonical_name",
            "match_type",
            "status",
            "notas",
        ],
        brand_alias,
    )
    write_csv(
        "normalization_dict.csv",
        [
            "norm_id",
            "domain",
            "raw_variant",
            "canonical_key",
            "canonical_display",
            "status",
            "notas",
        ],
        norm_dict,
    )
    write_csv(
        "category_bridge_candidates.csv",
        [
            "bridge_id",
            "source_system",
            "source_category_path",
            "suggested_official_codigo",
            "suggested_official_path",
            "oracle_path_exact_match",
            "internal_erp_codigo_hint",
            "match_method",
            "status",
            "notas",
        ],
        cat_bridge,
    )
    write_csv(
        "unit_master_candidates.csv",
        [
            "unit_candidate_id",
            "raw_presentation",
            "source",
            "parsed_quantity",
            "parsed_unit_token",
            "suggested_unity_id",
            "factor_to_base",
            "status",
            "notas",
        ],
        unit_rows,
    )

    summary_path = OUT / "RESUMEN_AMBIGUOS.txt"
    with summary_path.open("w", encoding="utf-8") as f:
        f.write("Resumen — normalización de catálogo (candidatos)\n")
        f.write("=" * 60 + "\n\n")
        f.write("Advertencias / archivos:\n")
        for w in warnings:
            f.write(f"  - {w}\n")
        f.write("\nConteos:\n")
        f.write(f"  brand_master: {len(brand_master)}\n")
        f.write(f"  brand_alias: {len(brand_alias)}\n")
        f.write(f"  normalization_dict: {len(norm_dict)}\n")
        f.write(f"  category_bridge_candidates: {len(cat_bridge)}\n")
        f.write(f"  unit_master_candidates: {len(unit_rows)}\n")
        f.write(f"  categorías únicas ecoplast: {len(eco_cats)}\n")
        f.write(f"  presentaciones únicas ecoplast: {len(eco_pres)}\n")
        f.write("\nAmbigüedades y pendientes (no auto-aprobados):\n")
        seen = set()
        for a in ambiguos:
            if a not in seen:
                seen.add(a)
                f.write(f"  - {a}\n")
        f.write("\nCriterios aplicados:\n")
        f.write("  - status=aprobado: solo match exacto contra catálogos JSON estructurados ")
        f.write("(nombre/abreviatura) o alias idéntico al canónico.\n")
        f.write("  - Marcas: clusters por minúsculas/espacios; todo propuesto salvo alias trivial.\n")
        f.write("  - Categorías: sin Categories_Tienda_Linea.xlsx no hay codigo/path oficial sugerido.\n")

    print("Salida en:", OUT)
    for w in warnings:
        print("AVISO:", w)
    return 0


if __name__ == "__main__":
    sys.exit(main())
