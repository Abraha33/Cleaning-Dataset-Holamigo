"""
Pobla brand_master, brand_alias, ambiguous_brands y normalization_dict (solo field=brand)
desde brands.xlsx (primario), brands.json y marcas_oraculo.json.
No fusiona por similitud; solo equivalencia nk (sin espacios, minúsculas) a marca única en xlsx.
"""
from __future__ import annotations

import csv
import json
import re
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]

PATH_XLSX = ROOT / "brands.xlsx"
PATH_JSON = ROOT / "brands.json"
PATH_ORACLE = ROOT / "marcas_oraculo.json"

OUT_MASTER = ROOT / "data-normalization" / "masters" / "brand_master.csv"
OUT_ALIAS = ROOT / "data-normalization" / "masters" / "brand_alias.csv"
OUT_NORM = ROOT / "data-normalization" / "masters" / "normalization_dict.csv"
OUT_AMBIG = ROOT / "data-normalization" / "review" / "ambiguous_brands.csv"

EXPECTED = {
    "brand_master": "brand_id,brand_name_canonical,brand_slug,brand_code_legacy,has_logo,logo_filename,source_primary,source_refs,is_active,notes",
    "brand_alias": "alias_id,brand_id,alias_name,alias_slug,alias_type,source,usage_count,status",
    "normalization_dict": "norm_id,field,original_value,normalized_value,normalization_type,source,usage_count,confidence,status,last_reviewed_by,last_reviewed_at",
    "ambiguous_brands": "raw_value,possible_brand_id,possible_brand_name,reason,confidence,status",
}


def nk(s: str) -> str:
    return "".join(str(s).lower().split())


def slugify(text: str) -> str:
    s = unicodedata.normalize("NFD", str(text))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "x"


def read_xlsx_brands(path: Path) -> list[tuple[str, str, str | None]]:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("Se requiere openpyxl: pip install openpyxl") from e
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active
    header = [str(c.value).strip() if c.value else "" for c in next(sh.iter_rows(min_row=1, max_row=1))]
    # Columnas observadas: Nombre, Código, Imagen
    def col(name_variants: tuple[str, ...]) -> int:
        for i, h in enumerate(header):
            hnorm = h.strip().lower()
            for v in name_variants:
                if hnorm == v.lower():
                    return i
        return -1

    i_nom = col(("Nombre", "nombre", "NOMBRE"))
    i_cod = col(("Código", "Codigo", "codigo", "CÓDIGO"))
    i_img = col(("Imagen", "imagen", "IMAGEN"))
    if i_nom < 0 or i_cod < 0:
        raise RuntimeError(f"brands.xlsx: encabezados no reconocidos. Obtenido: {header}")
    out = []
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or row[i_nom] is None:
            continue
        nombre = str(row[i_nom]).strip()
        if not nombre:
            continue
        cod = str(row[i_cod]).strip() if row[i_cod] is not None else ""
        img = None
        if i_img >= 0 and row[i_img] is not None:
            img = str(row[i_img]).strip() or None
        out.append((nombre, cod, img))
    wb.close()
    return out


def verify_headers() -> None:
    files = [
        (OUT_MASTER, "brand_master"),
        (OUT_ALIAS, "brand_alias"),
        (OUT_NORM, "normalization_dict"),
        (OUT_AMBIG, "ambiguous_brands"),
    ]
    for path, key in files:
        if not path.exists():
            raise FileNotFoundError(path)
        line = path.read_text(encoding="utf-8-sig").splitlines()[0].strip()
        exp = EXPECTED[key]
        if line != exp:
            raise ValueError(f"Header mismatch {path.name}:\n  got:     {line}\n  expected:{exp}")


def main() -> int:
    verify_headers()

    if not PATH_JSON.exists():
        raise FileNotFoundError(PATH_JSON)
    if not PATH_ORACLE.exists():
        raise FileNotFoundError(PATH_ORACLE)
    if not PATH_XLSX.exists():
        raise FileNotFoundError(PATH_XLSX)

    xlsx_rows = read_xlsx_brands(PATH_XLSX)
    by_nk: dict[str, tuple[str, str, str | None]] = {}
    for nombre, cod, img in xlsx_rows:
        k = nk(nombre)
        if k in by_nk:
            raise RuntimeError(
                f"brands.xlsx: dos filas comparten clave nk={k!r} ({by_nk[k][0]!r} vs {nombre!r}); detener."
            )
        by_nk[k] = (nombre, cod, img)

    brands_json: list[str] = json.loads(PATH_JSON.read_text(encoding="utf-8"))
    oracle: list[str] = json.loads(PATH_ORACLE.read_text(encoding="utf-8"))

    json_nk_set = {nk(s) for s in brands_json}

    master_rows: list[dict] = []
    for nombre, cod, img in xlsx_rows:
        if not cod:
            raise RuntimeError(f"Fila sin código en brands.xlsx para nombre={nombre!r}")
        has_logo = bool(img)
        in_json = nk(nombre) in json_nk_set
        refs = "brands.xlsx,brands.json" if in_json else "brands.xlsx"
        master_rows.append(
            {
                "brand_id": cod,
                "brand_name_canonical": nombre,
                "brand_slug": slugify(nombre),
                "brand_code_legacy": "",
                "has_logo": "true" if has_logo else "false",
                "logo_filename": img or "",
                "source_primary": "brands.xlsx",
                "source_refs": refs,
                "is_active": "true",
                "notes": "",
            }
        )

    alias_rows: list[dict] = []
    norm_rows: list[dict] = []
    ambig_rows: list[dict] = []

    alias_seq = 0
    norm_seq = 0
    amb_seq = 0

    seen_alias: set[tuple[str, str]] = set()

    def add_alias(brand_id: str, alias_name: str, alias_type: str, source: str) -> None:
        nonlocal alias_seq
        key = (brand_id, alias_name)
        if key in seen_alias:
            return
        if alias_name == next((m["brand_name_canonical"] for m in master_rows if m["brand_id"] == brand_id), None):
            return
        alias_seq += 1
        seen_alias.add(key)
        alias_rows.append(
            {
                "alias_id": f"AL{alias_seq:05d}",
                "brand_id": brand_id,
                "alias_name": alias_name,
                "alias_slug": slugify(alias_name),
                "alias_type": alias_type,
                "source": source,
                "usage_count": "",
                "status": "propuesto",
            }
        )

    def add_norm(original: str, normalized: str) -> None:
        nonlocal norm_seq
        norm_seq += 1
        norm_rows.append(
            {
                "norm_id": f"ND{norm_seq:05d}",
                "field": "brand",
                "original_value": original,
                "normalized_value": normalized,
                "normalization_type": "strict_nk_single_xlsx",
                "source": "marcas_oraculo.json",
                "usage_count": "",
                "confidence": "",
                "status": "propuesto",
                "last_reviewed_by": "",
                "last_reviewed_at": "",
            }
        )

    def add_ambig(raw: str, pid: str, pname: str, reason: str) -> None:
        nonlocal amb_seq
        amb_seq += 1
        ambig_rows.append(
            {
                "raw_value": raw,
                "possible_brand_id": pid,
                "possible_brand_name": pname,
                "reason": reason,
                "confidence": "",
                "status": "propuesto",
            }
        )

    # marcas_oraculo: match nk a catálogo xlsx único
    for o in oracle:
        o = str(o).strip()
        if not o:
            continue
        if o == "Bambu / Chevere":
            add_ambig(
                o,
                "",
                "",
                "Cadena compuesta con dos marcas presentes en brands.xlsx (BAMBU código BAM, CHEVERE código CHE); "
                "sin regla explícita de mapeo a una sola marca.",
            )
            continue

        kn = nk(o)
        hit = by_nk.get(kn)
        if hit is None:
            add_ambig(
                o,
                "",
                "",
                "Sin coincidencia por clave nk (minúsculas y sin espacios) con ningún Nombre único en brands.xlsx; "
                "no se asigna marca automáticamente.",
            )
            continue

        canon, bid, _ = hit
        if o == canon:
            continue

        if canon.lower() == o.lower():
            atype = "case_variant"
        elif nk(o) == nk(canon):
            atype = "display_variant"
        else:
            atype = "display_variant"

        add_alias(bid, o, atype, "marcas_oraculo.json")
        add_norm(o, canon)

    # Mantener solo filas normalization_dict únicas por (original, normalized)
    uniq_norm: dict[tuple[str, str], dict] = {}
    for r in norm_rows:
        uniq_norm[(r["original_value"], r["normalized_value"])] = r
    norm_rows = list(uniq_norm.values())
    norm_rows.sort(key=lambda x: x["norm_id"])

    # Re-id norm after dedup
    for i, r in enumerate(norm_rows, start=1):
        r["norm_id"] = f"ND{i:05d}"

    # Escribir CSV (UTF-8 BOM)
    def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)

    write_csv(
        OUT_MASTER,
        EXPECTED["brand_master"].split(","),
        master_rows,
    )
    write_csv(
        OUT_ALIAS,
        EXPECTED["brand_alias"].split(","),
        alias_rows,
    )
    write_csv(
        OUT_NORM,
        EXPECTED["normalization_dict"].split(","),
        norm_rows,
    )
    write_csv(
        OUT_AMBIG,
        EXPECTED["ambiguous_brands"].split(","),
        ambig_rows,
    )

    print("OK", len(master_rows), len(alias_rows), len(norm_rows), len(ambig_rows))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as e:
        print("ERROR:", type(e).__name__, e, file=sys.stderr)
        sys.exit(1)
